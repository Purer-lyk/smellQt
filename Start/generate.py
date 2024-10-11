import glob
import matplotlib.pyplot as plt
import torch
from torch.utils import data
import numpy as np
from matplotlib.pylab import style
from openpyxl import load_workbook

# 画图标题显示中文
style.use('ggplot')
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

ori_path = "./data/"
train_path = ["train.xlsx"]
test_path = ["test.xlsx"]

label_list = ["c2h5oh", "ch2oh", "thin", "strong"]
mtype_list = ['Attention', 'CNN', 'MLP']


class cnn_dataset(data.Dataset):
    def __init__(self, cnn_input_tensor, input_label):
        super(cnn_dataset, self).__init__()
        self.data_tensor = cnn_input_tensor
        self.input_label = input_label

    def __len__(self):
        return len(self.data_tensor)

    def __getitem__(self, item):
        single_tensor = self.data_tensor[item]
        single_tensor = np.reshape(single_tensor, [1, 10, 10])
        single_label = self.input_label[item]
        return single_tensor, single_label


class transformer_dataset(data.Dataset):
    def __init__(self, input_tensor, input_labels):
        super(transformer_dataset, self).__init__()
        self.data_tensor = input_tensor
        self.get_labels = input_labels

    def __len__(self):
        return len(self.data_tensor)

    def __getitem__(self, item):
        single_tensor = self.data_tensor[item]
        image_tensor = np.int32(np.round(single_tensor * 255.0))
        # print(image_tensor)
        single_label = self.get_labels[item]
        return image_tensor, single_label


class mlp_dataset(data.Dataset):
    def __init__(self, input_tensor, input_labels):
        super(mlp_dataset, self).__init__()
        self.data_tensor = input_tensor
        self.get_labels = input_labels

    def __len__(self):
        return len(self.data_tensor)

    def __getitem__(self, item):
        single_tensor = self.data_tensor[item]
        single_label = self.get_labels[item]
        return single_tensor, single_label


def list_init(xls):
    train_data = []
    train_label = []

    xls_book = load_workbook(ori_path + xls)
    sheets = xls_book.worksheets

    need_sheet = sheets[0]
    rows = need_sheet.rows

    k = 0
    for row in rows:
        if not k:
            k += 1
            continue
        row_val = [cell.value for cell in row]
        train_data.append(row_val[1:])
        train_label.append(row_val[0])
    return train_data, train_label


def generate_nums(xls_index):
    train_data, train_label = list_init(xls_index)  # 训练数据初始化

    data_array = np.array(train_data, dtype=np.float32)
    label_array = np.array(train_label, dtype=np.int32) - 1
    ########
    # plt.figure(num=3, figsize=(8, 5))
    # plt.title("乙醇0和甲醇0")
    # plt.ylabel("归一化后的响应")
    # plt.xlabel("时间")
    # print(len(data_array[0]))
    # x = range(100)
    # # for i in range(len(data_array)):
    # #     plt.plot(x, data_array[i])
    # plt.plot(x, data_array[0])
    # plt.plot(x, data_array[50])
    # print(label_array[0], label_array[50])
    # plt.show()
    ########
    # print(data_array)
    # print(label_array)
    return data_array, label_array


def experiment(xls):
    xls_book = load_workbook(ori_path + xls)
    sheets = xls_book.worksheets

    need_sheet = sheets[1]
    print(need_sheet.title)
    cols = need_sheet.columns

    for j, col in enumerate(cols):
        if j == 4 or j == 8:
            plt.figure(num=3, figsize=(8, 5))
            plt.title("差分数据")
            plt.ylabel("归一化后的响应")
            plt.xlabel("时间")
            x = range(99)
            col_val = [cell.value for cell in col]
            print(col_val)
            plt.plot(x, col_val)
            plt.show()


if __name__ == "__main__":
    # generate_nums(train_path[0])
    experiment("train_diff.xlsx")
